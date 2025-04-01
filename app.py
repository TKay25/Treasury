from flask import Flask, request, jsonify, session, render_template, redirect, url_for, send_file,flash
import psycopg2
from psycopg2 import pool
import pandas as pd
from datetime import timedelta
import uuid
import openpyxl
import numpy as np
from datetime import datetime, timedelta
import re
from io import BytesIO, StringIO
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.drawing.image import Image
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment



# Flask app setup
app = Flask(__name__)
app.secret_key = '011235811'
app.permanent_session_lifetime = timedelta(minutes=30)
user_sessions = {}
today_date = datetime.now().strftime('%d %B %Y')

# PostgreSQL Database Connection Pool
DATABASE_URL = "postgresql://treasury_l7wr_user:MVeyJTgGFi3ZzdRjx6Gy3byLqv0aqEKz@dpg-cv473trqf0us73b6hveg-a.oregon-postgres.render.com/treasury_l7wr"

try:
    db_pool = psycopg2.pool.SimpleConnectionPool(1, 20, dsn=DATABASE_URL)
    if db_pool:
        print("✅ Connection pool created successfully!")

        def get_db_connection():
            """Get a database connection from the pool."""
            try:
                return db_pool.getconn()
            except Exception as e:
                print("🚨 Error getting DB connection:", e)
                return None

        def release_db_connection(conn):
            """Release a connection back to the pool."""
            if conn:
                db_pool.putconn(conn)

        def delete_all_tables():
            try:
                conn = get_db_connection()
                cursor = conn.cursor()

                
                # Get all table names in the public schema
                cursor.execute("""
                    SELECT tablename FROM pg_tables WHERE schemaname = 'public';
                """)
                tables = cursor.fetchall()

                # Loop through and drop each table
                for table in tables:
                    table_name = table[0]
                    cursor.execute(f"DROP TABLE IF EXISTS {table_name} CASCADE;")
                    print(f"Dropped table: {table_name}")
                    conn.commit()

                # Close the cursor and connection
                print("All tables deleted successfully.")

            except Exception as e:
                print("Error:", e)

        # Run the function
        #delete_all_tables()
        table_name_users = "userstmo"
        table_name_action_items = "actionitemstmo"
        table_name_mm = "tmomm"
        table_name_mm_deleted = "tmommdeleted"

        def create_table_init():
            conn = get_db_connection()
            cursor = conn.cursor()
            create_table_query = f"""CREATE TABLE {table_name_users} (id SERIAL PRIMARY KEY,username VARCHAR(100),password VARCHAR(100));"""
            cursor.execute(create_table_query)
            conn.commit()
            print("Table USERS created successfully!")

        #create_table_init()

        def create_table_tmomm():
            conn = get_db_connection()
            cursor = conn.cursor()
            create_table_query = f"""CREATE TABLE {table_name_mm_deleted} (CALid SERIAL PRIMARY KEY, CALLoggerid INT, CALLogger VARCHAR(100), MarketCategory VARCHAR(100),	DealReference VARCHAR(100),	DealType VARCHAR(100), Counterparty VARCHAR(100),	Currency VARCHAR(100),	EffectedDate date, ValueDate date,	DaysDelay INT,	KnockoffCALid INT,	SupposedApprover VARCHAR(100),	Approver VARCHAR(100));"""
            cursor.execute(create_table_query)
            conn.commit()
            print("Table USERS created successfully!")

        #create_table_tmomm()

        def create_table():
            conn = get_db_connection()
            cursor = conn.cursor()
            create_table_query = f"""CREATE TABLE {table_name_action_items} (taskid SERIAL PRIMARY KEY,tasktitle VARCHAR(100),taskowner VARCHAR(100), priority VARCHAR(100), source VARCHAR(100), startdate date, duedate date, progressperc INT, status VARCHAR(100), statusdate date, comments VARCHAR(200), commentsdate date);"""
            cursor.execute(create_table_query)
            conn.commit()
            print("Table action items created successfully!")

        #create_table()

        def add_column():
            conn = get_db_connection()
            add_column_query = f"""ALTER TABLE {table_name_mm_deleted} ADD COLUMN comments VARCHAR(200);"""
            cursor = conn.cursor()
            cursor.execute(add_column_query)
            conn.commit()
            print("Column added successfully!")
        #add_column()

        def create_user(username, password):
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                insert_query = f"""INSERT INTO {table_name_users} (username, password) 
                                VALUES (%s, %s);"""
                cursor.execute(insert_query, (username, password))
                conn.commit()
                print(f"User {username} inserted successfully.")
            except Exception as e:
                print("Error inserting user:", e)


        #create_user('rmabika','55337')

        def run1(empid):
            global today_date

            global table_name_mm
            conn = get_db_connection()
            cursor = conn.cursor()
            query = f"SELECT CALid, CALLoggerid, CALLogger, TO_CHAR( DateLogged, 'FMDD Month YYYY') AS DateLogged,Market,  MarketCategory, DealReference, DealType, Currency, TO_CHAR( EffectedDate, 'FMDD Month YYYY') AS EffectedDate,  TO_CHAR( ValueDate, 'FMDD Month YYYY') AS ValueDate , Counterparty, DaysDelay, Approver, SupposedApprover, KnockoffCALid, Count, comments FROM {table_name_mm};"
            cursor.execute(query)
            rows = cursor.fetchall()

            mmlog = pd.DataFrame(rows, columns=["CAL ID","CAL LOGGER ID", "LOGGER", "LOGGED ON", "MARKET","AMENDMENT","DEAL REFERENCE", "DEAL TYPE", "CURRENCY", "EFFECTEDDATE","VALUEDATE","COUNTERPARTY","DAYS DELAYED","APPROVER","SUPPOSED APPROVER", "KNOCKOFF CAL ID","COUNT","COMMENTS"])
            print(mmlog)
            mmlog['ACTION'] = mmlog['CAL ID'].apply(
                lambda x: f'''<div style="display: flex; gap: 10px;font-size: 12px;"> <button class="btn btn-primary3 view-cal-btn" style="font-weight:bolder;" data-bs-toggle="modal" data-bs-target="#viewcalidModal" data-name="{x}"  data-ID="{x}">VIEW</button> </div>'''
            )
            mmlogall = mmlog[["CAL ID","CAL LOGGER ID", "LOGGER", "LOGGED ON", "MARKET","AMENDMENT","DEAL REFERENCE", "DEAL TYPE", "CURRENCY", "EFFECTEDDATE","VALUEDATE","COUNTERPARTY","DAYS DELAYED","APPROVER","SUPPOSED APPROVER", "KNOCKOFF CAL ID","COUNT","COMMENTS","ACTION"]]

            mycallog = mmlog[mmlog['CAL LOGGER ID'] == empid]
            mycallog['ACTION'] = mycallog['CAL ID'].apply(
                lambda x: f'''<div style="display: flex; gap: 10px;font-size: 12px;"> <button class="btn btn-primary3 edit-cal-btn" style="font-weight:bolder;"  data-bs-toggle="modal" data-bs-target="#editModalcal" data-name="{x}"  data-ID="{x}">EDIT</button>  <button class="btn btn-primary3 delete-cal-btn" style="font-weight:bolder;" data-bs-toggle="modal" data-bs-target="#deleteModalcal" data-name="{x}" data-ID="{x}">DELETE</button> </div>'''
            )
            mycallog = mycallog[["CAL ID","CAL LOGGER ID", "LOGGER", "LOGGED ON", "MARKET","AMENDMENT","DEAL REFERENCE", "DEAL TYPE", "CURRENCY", "EFFECTEDDATE","VALUEDATE","COUNTERPARTY","DAYS DELAYED","APPROVER","SUPPOSED APPROVER", "KNOCKOFF CAL ID","COUNT", "COMMENTS", "ACTION"]]

            mmlogall = mmlogall.fillna("")
            mycallog = mycallog.fillna("")

            table_mm_log_html = mmlogall.to_html(classes="table table-bordered table-theme", table_id="mmlogtable", index=False,  escape=False,)
            table_my_mm_log_html = mycallog.to_html(classes="table table-bordered table-theme", table_id="mymmlogtable", index=False,  escape=False,)


            return {
                "today_date": today_date,
                "table_mm_log_html": table_mm_log_html,
                "table_my_mm_log_html": table_my_mm_log_html,


            }




        TABLE_NAME_USERS = "userstmo"





        @app.route('/CAL_log', methods=['POST'])
        def callog():
            holidays = ['2025-02-21', '2025-04-18']  # Example holidays

            user_uuid = session.get('user_uuid')
            if not user_uuid:
                return redirect(url_for('landingpage'))  # Ensure a response is returned

            username = session.get('username')
            empid = session.get('empid')

            # Retrieve form data
            market = request.form.get('market')
            calcatmm = request.form.get('calcatmm')
            deal_reference = request.form.get('mm-deal-reference')
            capturedelay_options = request.form.get('capturedelayOptions')
            termination_capture_date = request.form.get('delayedterminationcaptureDate')
            termination_value_date = request.form.get('delayedterminationvalueDate')
            cancellation_capture_date = request.form.get('delayedcancellationDate')
            cancellation_value_date = request.form.get('supposedcancellationDate')
            backdatedcapturevaluedate = request.form.get('valueDate')
            unmatureddealvaluedate = request.form.get('supposedmaturityDate')
            approver = request.form.get('approveroutofofficeofficerOptions')
            supposed_approver = request.form.get('outofofficeofficerOptions')
            supposedoutofofficecapturedate = request.form.get('supposedoutofofficecaptureDate')
            dealcount = request.form.get('dealcount')
            uncapturedsupposedcapture = request.form.get('supposedcaptureDate')

            print(cancellation_capture_date)
            print(cancellation_value_date)

            date_entered = request.form.get('dateentered')
            principal_change_options = request.form.get('principalChangeOptions')
            mm_deal_type = request.form.get('uncaptureddealtypeOptions')
            print("ahhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh")
            print(mm_deal_type)
            currency = request.form.get('ccyOptions')
            interest_change_options = request.form.get('interestrateChangeOptions')
            fx_deal_bank_name = request.form.get('fxdealbankname')
            counterparty_mm = request.form.get('counterpartymm')
            buy_currency = request.form.get('calcatfxbuycurrency')
            sell_currency = request.form.get('calcatfxsellcurrency')
            fx_deal_date = request.form.get('fxdealDate')
            fx_deal_approval_date = request.form.get('fxdealapprovalDate')
            minimum_supposed_fx_approver = request.form.get('minimumsupposedfxapproverOptions')
            approver_out_of_office = request.form.get('approveroutofofficeofficerOptions')
            comments = request.form.get('comments')
            knockoffid = request.form.get('knockoffid')

            # Validate CAL Category
            if calcatmm == 'Early Termination':
                deal_type_mapping = {
                    'nncd': "NNCD",
                    'inbr': "Treasury Bill",
                    'inop': "Offshore Placement",
                    'inpp': "Local Placement",
                    'fixd': "Fixed Deposit",
                    'inbd': "Interbank Deposit",
                    'cctd': "Cash Cover Term Deposit",
                    'repr': "Repurchase Agreement",

                }

                # Extract the first 4 characters of deal_reference and convert to lowercase
                termination_deal_type = deal_reference[:4].lower()

                # Get the deal_type from the dictionary, default to "Unknown Deal Type" if not found
                deal_type = deal_type_mapping.get(termination_deal_type, "Unknown Deal Type")

                if "za" in deal_reference.lower():
                    currency = "ZAR"
                elif "eu" in deal_reference.lower():
                    currency = "EUR"
                elif "zg" in deal_reference.lower():
                    currency = "ZWG"
                else:
                    currency = "USD"  

                working_days_count = 0

                if termination_value_date and termination_capture_date:
                    try:
                        # Convert string dates to datetime objects
                        termination_value_date = datetime.strptime(termination_value_date, '%Y-%m-%d')
                        termination_capture_date = datetime.strptime(termination_capture_date, '%Y-%m-%d')

                        current_date = termination_value_date
                        while current_date <= termination_capture_date:
                            # Check if the current day is a weekday and not a holiday
                            if current_date.weekday() < 5 and current_date not in holidays:
                                working_days_count += 1
                            # Move to the next day
                            current_date += timedelta(days=1)
                    except Exception as e:
                        print("Error calculating working days:", e)
                        return redirect(url_for('landingpage'))  # Handle date parsing errors

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    if knockoffid:
                        if working_days_count == 0 :
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, DaysDelay, KnockoffCALid, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, working_days_count, knockoffid, comments))
                        else:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, KnockoffCALid, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, termination_capture_date, termination_value_date, working_days_count-1 , knockoffid, comments))
                    else:
                        if working_days_count == 0:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, DaysDelay, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, working_days_count, comments))
                        else:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, termination_capture_date, termination_value_date, working_days_count-1, comments))

                    conn.commit()
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

                except Exception as e:
                    print("Error inserting user:", e)
                    return redirect(url_for('landingpage'))  # Ensure a response is returned
                
            elif calcatmm == 'Uncaptured Deal':
                print(f"deal type : {mm_deal_type}")
                if mm_deal_type == "NNCD":
                    counterparty_mm = "RBZ"
                    currency = "ZWG"

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    dealcount = int(dealcount)
                    insert_query = f"""
                    INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealType, Currency, ValueDate, Count, Counterparty, comments)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, mm_deal_type, currency, uncapturedsupposedcapture, dealcount, counterparty_mm, comments))

                    conn.commit()
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

                except Exception as e:
                    print("Error inserting user:", e)
                    return redirect(url_for('landingpage'))  # Ensure a response is returned
                




            elif calcatmm == 'Cancelled Deal':
                deal_type_mapping = {
                    'nncd': "NNCD",
                    'inbr': "Treasury Bill",
                    'inop': "Offshore Placement",
                    'inpp': "Local Placement",
                    'fixd': "Fixed Deposit",
                    'inbd': "Interbank Deposit",
                    'cctd': "Cash Cover Term Deposit",
                }

                # Extract the first 4 characters of deal_reference and convert to lowercase
                termination_deal_type = deal_reference[:4].lower()

                # Get the deal_type from the dictionary, default to "Unknown Deal Type" if not found
                deal_type = deal_type_mapping.get(termination_deal_type, "Unknown Deal Type")

                if "za" in deal_reference.lower():
                    currency = "ZAR"
                elif "eu" in deal_reference.lower():
                    currency = "EUR"
                elif "zg" in deal_reference.lower():
                    currency = "ZWG"
                else:
                    currency = "USD"  

                working_days_count = 0

                if cancellation_value_date and cancellation_capture_date:
                    try:
                        # Convert string dates to datetime objects
                        cancellation_value_date = datetime.strptime(cancellation_value_date, '%Y-%m-%d')
                        cancellation_capture_date = datetime.strptime(cancellation_capture_date, '%Y-%m-%d')

                        current_date = cancellation_value_date
                        while current_date <= cancellation_capture_date:
                            # Check if the current day is a weekday and not a holiday
                            if current_date.weekday() < 5 and current_date not in holidays:
                                working_days_count += 1
                            # Move to the next day
                            current_date += timedelta(days=1)

                    except Exception as e:
                        print("Error calculating working days:", e)
                        return redirect(url_for('dashboard'))  # Handle date parsing errors

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    if working_days_count == 0:
                        insert_query = f"""
                        INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, DaysDelay, KnockoffCALid, comments)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        """
                        cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, working_days_count, knockoffid, comments))
                    else:
                        insert_query = f"""
                        INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, KnockoffCALid, comments)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        """
                        cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, cancellation_capture_date, cancellation_value_date, working_days_count-1, knockoffid, comments))

                    conn.commit()
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

                except Exception as e:
                    print("Error inserting user:", e)
                    return redirect(url_for('landingpage'))  # Ensure a response is returned
                


            elif calcatmm == 'Unmatured Deal':
                deal_type_mapping = {
                    'nncd': "NNCD",
                    'inbr': "Treasury Bill",
                    'inop': "Offshore Placement",
                    'inpp': "Local Placement",
                    'fixd': "Fixed Deposit",
                    'inbd': "Interbank Deposit",
                    'cctd': "Cash Cover Term Deposit",
                }

                # Extract the first 4 characters of deal_reference and convert to lowercase
                termination_deal_type = deal_reference[:4].lower()

                # Get the deal_type from the dictionary, default to "Unknown Deal Type" if not found
                deal_type = deal_type_mapping.get(termination_deal_type, "Unknown Deal Type")

                if "za" in deal_reference.lower():
                    currency = "ZAR"
                elif "eu" in deal_reference.lower():
                    currency = "EUR"
                elif "zg" in deal_reference.lower():
                    currency = "ZWG"
                else:
                    currency = "USD"  


                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    insert_query = f"""
                    INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, ValueDate, comments)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, unmatureddealvaluedate, comments))

                    conn.commit()
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

                except Exception as e:
                    print("Error inserting user:", e)
                    return redirect(url_for('landingpage'))  # Ensure a response is returned
                
                
            elif calcatmm == 'Out of Office TFO':
                deal_type_mapping = {
                    'nncd': "NNCD",
                    'inbr': "Treasury Bill",
                    'inop': "Offshore Placement",
                    'inpp': "Local Placement",
                    'fixd': "Fixed Deposit",
                    'inbd': "Interbank Deposit",
                    'cctd': "Cash Cover Term Deposit",
                }

                # Extract the first 4 characters of deal_reference and convert to lowercase
                termination_deal_type = deal_reference[:4].lower()

                # Get the deal_type from the dictionary, default to "Unknown Deal Type" if not found
                deal_type = deal_type_mapping.get(termination_deal_type, "Unknown Deal Type")

                if "za" in deal_reference.lower():
                    currency = "ZAR"
                elif "eu" in deal_reference.lower():
                    currency = "EUR"
                elif "zg" in deal_reference.lower():
                    currency = "ZWG"
                else:
                    currency = "USD"  

                number_part = re.search(r'\d+', deal_reference)
                print('here')
                numberpartextract = number_part.group()
                print(numberpartextract)
                datepart = numberpartextract[:-1]
                print(datepart)


                number_str = str(datepart)

                # Determine the length of the number
                length = len(number_str)

                # Extract day, month, and year based on the length
                if length == 7:
                    # Assuming format is MMDDYYYY
                    day = int(number_str[:1])  # First digit is month
                    month = int(number_str[1:3])   # Next two digits are day
                    year = int(number_str[3:])   # Remaining digits are year
                elif length == 8:
                    # Assuming format is MMDDYYYY or DDMMYYYY
                    day = int(number_str[:2])  # First two digits are month
                    month = int(number_str[2:4])   # Next two digits are day
                    year = int(number_str[4:])   # Remaining digits are year
                else:
                    raise ValueError("Invalid number length for date conversion")

                # Create a date object
                try:
                    backdated_capture_date = datetime(year, month, day).strftime('%Y-%m-%d')
                    print(f"Date: {backdated_capture_date}")

                except ValueError as e:
                    print(f"Invalid date: {e}")

                working_days_count = 0

                if supposedoutofofficecapturedate and backdated_capture_date:
                    try:
                        # Convert string dates to datetime objects
                        supposedoutofofficecapturedate = datetime.strptime(supposedoutofofficecapturedate, '%Y-%m-%d')
                        backdated_capture_date = datetime.strptime(backdated_capture_date, '%Y-%m-%d')

                        current_date = supposedoutofofficecapturedate
                        while current_date <= backdated_capture_date:
                            # Check if the current day is a weekday and not a holiday
                            if current_date.weekday() < 5 and current_date not in holidays:
                                working_days_count += 1
                            # Move to the next day
                            current_date += timedelta(days=1)

                    except Exception as e:
                        print("Error calculating working days:", e)
                        return redirect(url_for('dashboard'))  # Handle date parsing errors

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    if knockoffid:
                        if working_days_count == 0:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, DaysDelay, Approver, ValueDate, SupposedApprover, KnockoffCALid, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, working_days_count, approver, backdated_capture_date, supposed_approver, knockoffid, comments))
                        else:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, Approver, SupposedApprover,KnockoffCALid, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, backdated_capture_date, supposedoutofofficecapturedate, working_days_count-1, approver, supposed_approver, knockoffid, comments))
                    else:
                        if working_days_count == 0:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, DaysDelay, Approver, SupposedApprover, ValueDate, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, working_days_count, approver, supposed_approver, backdated_capture_date, comments))
                        else:
                            insert_query = f"""
                            INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, Approver, SupposedApprover, comments)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, backdated_capture_date, supposedoutofofficecapturedate, working_days_count-1, approver, supposed_approver, comments))

                    conn.commit()
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

                except Exception as e:
                    print("Error inserting user:", e)
                    return redirect(url_for('landingpage'))  # Ensure a response is returned
                

            elif calcatmm == 'Backdated Capture':
                deal_type_mapping = {
                    'nncd': "NNCD",
                    'inbr': "Treasury Bill",
                    'inop': "Offshore Placement",
                    'inpp': "Local Placement",
                    'fixd': "Fixed Deposit",
                    'inbd': "Interbank Deposit",
                    'cctd': "Cash Cover Term Deposit",
                }

                # Extract the first 4 characters of deal_reference and convert to lowercase
                termination_deal_type = deal_reference[:4].lower()

                # Get the deal_type from the dictionary, default to "Unknown Deal Type" if not found
                deal_type = deal_type_mapping.get(termination_deal_type, "Unknown Deal Type")

                if "za" in deal_reference.lower():
                    currency = "ZAR"
                elif "eu" in deal_reference.lower():
                    currency = "EUR"
                elif "zg" in deal_reference.lower():
                    currency = "ZWG"
                else:
                    currency = "USD"  

                number_part = re.search(r'\d+', deal_reference)
                print('here')
                numberpartextract = number_part.group()
                print(numberpartextract)
                datepart = numberpartextract[:-1]
                print(datepart)


                number_str = str(datepart)

                # Determine the length of the number
                length = len(number_str)

                # Extract day, month, and year based on the length
                if length == 7:
                    # Assuming format is MMDDYYYY
                    day = int(number_str[:1])  # First digit is month
                    month = int(number_str[1:3])   # Next two digits are day
                    year = int(number_str[3:])   # Remaining digits are year
                elif length == 8:
                    # Assuming format is MMDDYYYY or DDMMYYYY
                    day = int(number_str[:2])  # First two digits are month
                    month = int(number_str[2:4])   # Next two digits are day
                    year = int(number_str[4:])   # Remaining digits are year
                else:
                    raise ValueError("Invalid number length for date conversion")

                # Create a date object
                try:
                    backdated_capture_date = datetime(year, month, day).strftime('%Y-%m-%d')
                    print(f"Date: {backdated_capture_date}")

                except ValueError as e:
                    print(f"Invalid date: {e}")

                working_days_count = 0

                if backdatedcapturevaluedate and backdated_capture_date:
                    try:
                        # Convert string dates to datetime objects
                        backdatedcapturevaluedate = datetime.strptime(backdatedcapturevaluedate, '%Y-%m-%d')
                        backdated_capture_date = datetime.strptime(backdated_capture_date, '%Y-%m-%d')

                        current_date = backdatedcapturevaluedate
                        while current_date <= backdated_capture_date:
                            # Check if the current day is a weekday and not a holiday
                            if current_date.weekday() < 5 and current_date not in holidays:
                                working_days_count += 1
                            # Move to the next day
                            current_date += timedelta(days=1)

                    except Exception as e:
                        print("Error calculating working days:", e)
                        return redirect(url_for('dashboard'))  # Handle date parsing errors

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    if knockoffid:

                        insert_query = f"""
                        INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, KnockoffCALid, comments)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        """
                        cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, backdated_capture_date, backdatedcapturevaluedate, working_days_count-1, knockoffid, comments))

                    else:

                        insert_query = f"""
                        INSERT INTO {table_name_mm} (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, DealReference, DealType, Currency, EffectedDate, ValueDate, DaysDelay, comments)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        """
                        cursor.execute(insert_query, (today_date, empid, username, market, calcatmm, deal_reference, deal_type, currency, backdated_capture_date, backdatedcapturevaluedate, working_days_count-1, comments))

                    conn.commit()
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

                except Exception as e:
                    print("Error inserting user:", e)
                    return redirect(url_for('landingpage'))  # Ensure a response is returned\
                

        @app.route('/export_excel_alltime_deleted_callog')
        def export_deleted_excel():
            user_uuid = session.get('user_uuid')
            if user_uuid:

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    global table_name_mm
                    global today_date
                    global table_name_mm_deleted

                    query = f"SELECT * FROM {table_name_mm_deleted};"
                    cursor.execute(query)
                    rows = cursor.fetchall()
                    print(rows)

                    allmmcallog = pd.DataFrame(rows, columns=["CAL ID","CAL LOGGER ID","LOGGER","AMENDMENT","DEAL REFERENCE","DEAL TYPE","COUNTERPARTY","CURRENCY","EFFECTED DATE","VALUE DATE","DAYS DELAYED","KNOCKOFF ID","SUPPOSED APPROVER","APPROVER","MARKET","DATE LOGGED","COUNT","COMMENTS/NOTES"])
                    allmmcallog = allmmcallog[["CAL ID","CAL LOGGER ID","DATE LOGGED","LOGGER","MARKET","AMENDMENT","DEAL REFERENCE","DEAL TYPE","COUNTERPARTY","CURRENCY","EFFECTED DATE","VALUE DATE","DAYS DELAYED","KNOCKOFF ID","SUPPOSED APPROVER","APPROVER","COUNT","COMMENTS/NOTES"]]
                    
                    print(allmmcallog)

                    allmmcallog = allmmcallog.sort_values(by="CAL ID", ascending=False)

                    print(allmmcallog)

                    # Create an in-memory Excel file
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        allmmcallog.to_excel(writer, index=False, sheet_name=f'DELETED CAL {today_date}')

                    output.seek(0)
                    print('done')

                    # Send the file to the client
                    return send_file(
                        output,
                        as_attachment=True,
                        download_name=f'Deleted CAL Log Report as at {today_date}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                
                except Exception as e:
                    print("Error:", e)
                    return redirect(url_for('dashboard'))  # Ensure a response is returned

            else:
                return redirect(url_for('landingpage'))
     


        @app.route('/export_excel_alltime_callog')
        def export_excel():
            user_uuid = session.get('user_uuid')
            if user_uuid:
                global table_name_mm
                global today_date
                focus = request.args.get('priorityFocus')

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    query = f"SELECT * FROM {table_name_mm};"
                    cursor.execute(query)
                    rows = cursor.fetchall()

                    allmmcallog = pd.DataFrame(rows, columns=["CAL ID","CAL LOGGER ID","LOGGER","AMENDMENT","DEAL REFERENCE","DEAL TYPE","COUNTERPARTY","CURRENCY","EFFECTED DATE","VALUE DATE","DAYS DELAYED","KNOCKOFF ID","SUPPOSED APPROVER","APPROVER","MARKET","DATE LOGGED","COUNT","COMMENTS/NOTES"])
                    allmmcallog = allmmcallog[["CAL ID","CAL LOGGER ID","DATE LOGGED","LOGGER","MARKET","AMENDMENT","DEAL REFERENCE","DEAL TYPE","COUNTERPARTY","CURRENCY","EFFECTED DATE","VALUE DATE","DAYS DELAYED","KNOCKOFF ID","SUPPOSED APPROVER","APPROVER","COUNT","COMMENTS/NOTES"]]
                    
                    print(allmmcallog)

                    allmmcallog = allmmcallog.sort_values(by="CAL ID", ascending=False)
                    delayed = allmmcallog[allmmcallog['DAYS DELAYED'] > 0]
                    print(allmmcallog)

                    if focus == "alltime":

                        # Create an in-memory Excel file
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            allmmcallog.to_excel(writer, index=False, sheet_name=f'ALL CAL {today_date}')
                            delayed.to_excel(writer, index=False, sheet_name=f'DELAYED CAL {today_date}')

                        output.seek(0)
                        print('done')

                        # Send the file to the client
                        return send_file(
                            output,
                            as_attachment=True,
                            download_name=f'CAL Log Report as at {today_date}.xlsx',
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                    
                    elif focus == "customcalenders":

                        start_date = request.args.get('startDate')
                        end_date = request.args.get('endDate')

                        filtered_df = allmmcallog[
                            (allmmcallog['AMENDMENT'] == 'Unmatured Deal') & 
                            (pd.to_datetime(allmmcallog['VALUE DATE']) >= start_date) & 
                            (pd.to_datetime(allmmcallog['VALUE DATE']) <= end_date)
                        ]


                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            filtered_df.to_excel(writer, index=False, sheet_name=f'CAL REPORT {end_date}')

                        output.seek(0)
                        print('done')

                        # Send the file to the client
                        return send_file(
                            output,
                            as_attachment=True,
                            download_name=f'CAL Log Report for period {start_date} to {end_date}.xlsx',
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                    

                    elif focus == "daily":
                        # Extract the day from either VALUE DATE or EFFECTED DATE
                        def get_current_month_day(row):
                            # Check VALUE DATE first
                            val_date = pd.to_datetime(row['VALUE DATE'])
                            if val_date.month == current_month and val_date.year == current_year:
                                return val_date.day
                            
                            eff_date = pd.to_datetime(row['EFFECTED DATE'])
                            if eff_date.month == current_month and eff_date.year == current_year:
                                return eff_date.day
                            
                            return None  # Exclude dates outside current month
                        
                        asat_date = request.args.get('asatDate')

                        asat_date2 = pd.to_datetime(asat_date)

                        # Get previous business day
                        previous_business_day = asat_date2 - pd.offsets.BDay(1)

                        # If you need it as a string in the same format
                        previous_business_day_str = previous_business_day.strftime('%Y-%m-%d')


                        filtered_df1 = allmmcallog[
                            ((allmmcallog['AMENDMENT'] == 'Unmatured Deal') | (allmmcallog['AMENDMENT'] == 'Uncaptured Deal')) & 
                            ((pd.to_datetime(allmmcallog['VALUE DATE']) == asat_date) | (pd.to_datetime(allmmcallog['VALUE DATE']) == previous_business_day_str))
                        ]

                        filtered_df2 = allmmcallog[
                            ((allmmcallog['AMENDMENT'] != 'Unmatured Deal') | (allmmcallog['AMENDMENT'] != 'Uncaptured Deal')) & 
                            ((pd.to_datetime(allmmcallog['EFFECTED DATE']) == asat_date) | (pd.to_datetime(allmmcallog['VALUE DATE']) == previous_business_day_str))
                        ]

                        combined_df = filtered_df1._append(filtered_df2, ignore_index=True)

                        # Create filters for both dates
                        asat_date_filter = ((pd.to_datetime(allmmcallog['VALUE DATE']) == asat_date2) |(pd.to_datetime(allmmcallog['EFFECTED DATE']) == asat_date2))

                        prev_date_filter = ((pd.to_datetime(allmmcallog['VALUE DATE']) == previous_business_day) |(pd.to_datetime(allmmcallog['EFFECTED DATE']) == previous_business_day))

                        # Group by amendment and sum COUNT for each date
                        asat_date_totals = allmmcallog[asat_date_filter].groupby('AMENDMENT')['COUNT'].sum().reset_index()
                        prev_date_totals = allmmcallog[prev_date_filter].groupby('AMENDMENT')['COUNT'].sum().reset_index()

                        # Merge the totals
                        summary = pd.merge(
                            asat_date_totals.rename(columns={'COUNT': str(asat_date2)}),
                            prev_date_totals.rename(columns={'COUNT': str(previous_business_day)}),
                            on='AMENDMENT',
                            how='outer'
                        ).fillna(0)
                        # Calculate MTD (Month-to-Date) totals
                        current_month = asat_date2.month
                        current_year = asat_date2.year

                        mtd_filter = (
                            (pd.to_datetime(allmmcallog['VALUE DATE']).dt.month == current_month) &
                            (pd.to_datetime(allmmcallog['VALUE DATE']).dt.year == current_year)
                        ) | (
                            (pd.to_datetime(allmmcallog['EFFECTED DATE']).dt.month == current_month) &
                            (pd.to_datetime(allmmcallog['EFFECTED DATE']).dt.year == current_year)
                        )

                        mtd_totals = allmmcallog[mtd_filter].groupby('AMENDMENT')['COUNT'].sum().reset_index()

                        # Add MTD to summary
                        summary = pd.merge(
                            summary,
                            mtd_totals,
                            on='AMENDMENT',
                            how='left'
                        ).fillna(0)

                        # Rename columns and calculate totals
                        summary = summary.rename(columns={
                            'COUNT': 'MTD'
                        })

                        # Reorder columns
                        final_columns = [
                            'AMENDMENT',
                            f'{asat_date2}',
                            f'{previous_business_day}',
                            'MTD'
                        ]
                        summary = summary[final_columns]
                        summary = summary.sort_values(f'{asat_date2}', ascending=False)

                        asat_date_totals = allmmcallog[asat_date_filter].groupby(['AMENDMENT', 'DEAL TYPE'])['COUNT'].sum().reset_index()
                        prev_date_totals = allmmcallog[prev_date_filter].groupby(['AMENDMENT', 'DEAL TYPE'])['COUNT'].sum().reset_index()

                        # Create MTD totals by Amendment and Deal Type
                        mtd_filter = (
                            (pd.to_datetime(allmmcallog['VALUE DATE']).dt.month == current_month) &
                            (pd.to_datetime(allmmcallog['VALUE DATE']).dt.year == current_year)
                        ) | (
                            (pd.to_datetime(allmmcallog['EFFECTED DATE']).dt.month == current_month) &
                            (pd.to_datetime(allmmcallog['EFFECTED DATE']).dt.year == current_year)
                        )
                        mtd_totals = allmmcallog[mtd_filter].groupby(['AMENDMENT', 'DEAL TYPE'])['COUNT'].sum().reset_index()

                        # Merge all totals
                        summary_by_dealtype = pd.merge(
                            asat_date_totals.rename(columns={'COUNT': str(asat_date2)}),
                            prev_date_totals.rename(columns={'COUNT': str(previous_business_day)}),
                            on=['AMENDMENT', 'DEAL TYPE'],
                            how='outer'
                        ).fillna(0)

                        summary_by_dealtype = pd.merge(
                            summary_by_dealtype,
                            mtd_totals.rename(columns={'COUNT': 'MTD'}),
                            on=['AMENDMENT', 'DEAL TYPE'],
                            how='left'
                        ).fillna(0)

                        # Reorder columns
                        final_columns = [
                            'AMENDMENT', 
                            'DEAL TYPE',
                            str(asat_date2),
                            str(previous_business_day),
                            'MTD'
                        ]


                        summary_by_dealtype = summary_by_dealtype[final_columns]

                        # Sort by Amendment then by asat_date count
                        summary_by_dealtype = summary_by_dealtype.sort_values(
                            ['AMENDMENT', str(asat_date2)], 
                            ascending=[True, False]
                        )


                        def format_timestamp_column(col_name):
                            try:
                                # Parse timestamp (e.g., "2025-04-01 00:00:00")
                                dt = datetime.strptime(col_name, "%Y-%m-%d %H:%M:%S")
                                # Return formatted date (e.g., "01 April 2025")
                                return dt.strftime("%d %B %Y")  
                            except ValueError:
                                return col_name  # Return unchanged if not a timestamp

                        # Apply renaming
                        summary = summary.rename(columns=format_timestamp_column)
                        # Sort by TOTAL descending
                        print(summary)

                        month_data = allmmcallog[
                            (
                                (pd.to_datetime(allmmcallog['VALUE DATE']).dt.month == current_month) & 
                                (pd.to_datetime(allmmcallog['VALUE DATE']).dt.year == current_year
                            ) | (
                                (pd.to_datetime(allmmcallog['EFFECTED DATE']).dt.month == current_month) & 
                                (pd.to_datetime(allmmcallog['EFFECTED DATE']).dt.year == current_year
                            )))
                        ].copy()                        


                        month_data['DAY'] = month_data.apply(get_current_month_day, axis=1)

                        print(month_data)
                        # Group by Amendment and Day
                        daily_counts = month_data.groupby(['AMENDMENT', 'DAY'])['COUNT'].sum().reset_index()

                        # Create the visualization
                        plt.figure(figsize=(14, 7))

                        # Define your custom color palette
                        custom_palette = {
                            # Uncaptured Deal (red shades)
                            ('Uncaptured Deal'): '#8B0000',        # Darkest Red
                            ('Backdated Capture'): '#00008B',      # Dark Blue
                            ('Unmatured Deal'): '#808080',         # Grey
                            ('Other Amendment'): '#A9A9A9'       # Dark Grey
                        }

                        plt.figure(figsize=(12, 8))  # Increased height to accommodate table

                        # Create the bar plot
                        ax1 = plt.subplot2grid((10, 1), (0, 0), rowspan=6)
                        sns.barplot(
                            data=daily_counts,
                            x='DAY',
                            y='COUNT',
                            hue='AMENDMENT',
                            palette=custom_palette,
                            estimator=sum,
                            errorbar=None,
                            ax=ax1
                        )

                        # Format the plot
                        ax1.set_title(f'Cancelled, Amended, Late trades, Backdated and Limit breaches ({asat_date2.strftime("%d %B %Y")})', fontsize=10)
                        ax1.set_xlabel('Day of Month', fontsize=9)
                        ax1.set_ylabel('Total Count', fontsize=9)
                        ax1.tick_params(labelsize=8)
                        ax1.legend(
                            title='Amendment Types',
                            bbox_to_anchor=(1.02, 1),
                            loc='upper left',
                            fontsize=8
                        )

                        # Create table data - pivot to show amendments vs days
                        table_data = daily_counts.pivot_table(
                            index='AMENDMENT',
                            columns='DAY',
                            values='COUNT',
                            aggfunc='sum',
                            fill_value=0
                        )

                        # Add table below the plot
                        ax2 = plt.subplot2grid((10, 1), (7, 0), rowspan=3)
                        ax2.axis('off')  # Hide axes
                        table = ax2.table(
                            cellText=table_data.values,
                            rowLabels=table_data.index,
                            colLabels=table_data.columns,
                            cellLoc='center',
                            loc='center',
                            colWidths=[0.1]*len(table_data.columns))

                        table.auto_set_font_size(False)
                        table.set_fontsize(8)
                        table.scale(1, 1.2)  # Scale cell sizes

                        # Highlight header row
                        for (row, col), cell in table.get_celld().items():
                            if row == 0:
                                cell.set_text_props(weight='bold')
                                cell.set_facecolor('#003366')  # Dark blue header
                                cell.set_text_props(color='white')

                        plt.tight_layout()

                        # Save and export to Excel
                        img_buffer1 = BytesIO()
                        plt.savefig(img_buffer1, format='png', dpi=150, bbox_inches='tight')
                        plt.close()




                        summary_by_dealtype = summary_by_dealtype.rename(columns=format_timestamp_column)
                        # Sort by TOTAL descending
                        print(summary_by_dealtype)

                        mtd_data = allmmcallog[mtd_filter].copy()


                        




                        mtd_data['DAY'] = mtd_data.apply(get_current_month_day, axis=1)

                        # Drop rows where DAY is None (dates outside current month)
                        mtd_data = mtd_data.dropna(subset=['DAY'])

                        # Convert DAY to integer
                        mtd_data['DAY'] = mtd_data['DAY'].astype(int)

                        # Now group by AMENDMENT and DAY
                        # Group by Amendment, Deal Type, and Day
                        daily_counts = mtd_data.groupby(['AMENDMENT', 'DEAL TYPE', 'DAY'])['COUNT'].sum().reset_index()


                        # Create a combined label column for the legend (optional but helpful)
                        daily_counts['AMENDMENT_DEALTYPE'] = (
                            daily_counts['AMENDMENT'] + " (" + daily_counts['DEAL TYPE'] + ")"
                        )


                        plt.figure(figsize=(8, 4))

                        # Define color palette
                        color_palette = {
                            # Uncaptured Deal (red shades)
                            ('Uncaptured Deal', 'NNCD'): '#8B0000',        # Darkest Red
                            ('Uncaptured Deal', 'Treasury Bill'): '#DC143C',  # Crimson
                            ('Uncaptured Deal', 'FX Swap'): '#FF6347',     # Tomato
                            ('Uncaptured Deal', 'Other'): '#FFA07A',       # Light Salmon
                            
                            # Backdated Capture (blue shades)
                            ('Backdated Capture', 'NNCD'): '#00008B',      # Dark Blue
                            ('Backdated Capture', 'Treasury Bill'): '#1E90FF',  # Dodger Blue
                            ('Backdated Capture', 'FX Swap'): '#4682B4',   # Steel Blue
                            ('Backdated Capture', 'Other'): '#87CEFA',     # Light Sky Blue
                            
                            # Other amendments
                            ('Unmatured Deal', 'IRS'): '#808080',         # Grey
                            ('Other Amendment', 'Other'): '#A9A9A9'       # Dark Grey
                        }

                        # Get unique days and amendments
                        days = sorted(daily_counts['DAY'].unique())
                        amendments = daily_counts['AMENDMENT'].unique()

                        # Set up bar positions
                        bar_width = 0.8 / len(amendments)
                        x_positions = np.arange(len(days))

                        # Create legend handles and labels explicitly
                        legend_handles = []
                        legend_labels = []

                        # Plot each amendment
                        for i, amendment in enumerate(amendments):
                            amendment_data = daily_counts[daily_counts['AMENDMENT'] == amendment]
                            bottom = np.zeros(len(days))
                            
                            # Plot each deal type
                            for deal_type in daily_counts['DEAL TYPE'].unique():
                                subset = amendment_data[amendment_data['DEAL TYPE'] == deal_type]
                                if not subset.empty:
                                    counts = subset.groupby('DAY')['COUNT'].sum().reindex(days, fill_value=0)
                                    color = color_palette.get((amendment, deal_type), '#D3D3D3')
                                    
                                    # Create the bar plot
                                    bar = plt.bar(
                                        x=x_positions + i*bar_width,
                                        height=counts,
                                        bottom=bottom,
                                        width=bar_width,
                                        color=color
                                    )
                                    
                                    # Add to legend only once per combination
                                    if (amendment, deal_type) in color_palette:
                                        legend_handles.append(bar)
                                        legend_labels.append(f"{amendment} - {deal_type}")
                                    
                                    bottom += counts

                        # Formatting
                        plt.title('CAL BY DEAL TYPE')
                        plt.xlabel(f'DAY OF {asat_date2.strftime("%B %Y")}')
                        plt.ylabel('COUNT')
                        plt.xticks(x_positions + bar_width*(len(amendments)-1)/2, days)

                        # Create the legend with all amendment-deal type combinations
                        plt.legend(
                            handles=legend_handles,
                            labels=legend_labels,
                            title='AMENDMENT - DEAL TYPE',
                            bbox_to_anchor=(0.5, -0.2),  # 0.5 = center, -0.2 = below graph
                            loc='upper center',
                            fontsize=8,
                            ncol=3,  # Adjust number of columns as needed
                            frameon=True
                        )

                        # Adjust layout to make space for legend
                        plt.tight_layout(rect=[0, 0.1, 1, 1])  # rect=[left, bottom, right, top]

                        plt.grid(axis='y', alpha=0.3)
                        plt.tight_layout()

                        # Save and export to Excel
                        img_buffer = BytesIO()
                        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
                        plt.close()

                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            combined_df.to_excel(writer, index=False, sheet_name='CAL REPORT')
                            summary.to_excel(writer, index=False, sheet_name='CAL Dashboard')
                            # Write summary DataFrame starting at A6
                            summary_by_dealtype.to_excel(writer, index=False, sheet_name='CAL Dashboard', startrow=5, header=True)
                            
                            workbook = writer.book
                            worksheet = workbook['CAL Dashboard']

                            img = Image(img_buffer1)
                            img.width = 700
                            img.height = 400
                            img.anchor = 'F2'
                            worksheet.add_image(img)
                            
                            img = Image(img_buffer)
                            img.width = 700
                            img.height = 400
                            img.anchor = 'R2'
                            worksheet.add_image(img)

                            header_fmt = {
                                'fill': PatternFill(
                                    start_color='FF003366',  # Dark blue background (ARGB)
                                    end_color='FF003366',
                                    fill_type='solid'
                                ),
                                'font': Font(
                                    color='FFFFFFFF',  # White text (ARGB)
                                    bold=True,
                                    name='Calibri',
                                    size=11
                                ),
                                'border': Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                ),
                                'alignment': Alignment(
                                    horizontal='center',
                                    vertical='center'
                                )
                            }

                            # Apply to header row (row 6)
                            for col in range(1, len(summary.columns) + 1):
                                cell = worksheet.cell(row=1, column=col)
                                for attr, value in header_fmt.items():
                                    setattr(cell, attr, value)

                            for col in range(1, len(summary_by_dealtype.columns) + 1):
                                cell = worksheet.cell(row=6, column=col)
                                for attr, value in header_fmt.items():
                                    setattr(cell, attr, value)


                        output.seek(0)
                        return send_file(
                            output,
                            as_attachment=True,
                            download_name=f'CAL Log Report as at {asat_date}.xlsx',
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )



                    else:
                        return jsonify({'success': False, 'message': 'Invalid focus parameter'}), 400

                except Exception as e:
                    print("Error:", e)
                    return redirect(url_for('dashboard'))  # Ensure a response is returned
                    







            else:
                return redirect(url_for('landingpage'))


        @app.route('/delete_cal', methods=['POST'])
        def delete_cal():
            try:
                global table_name_mm
                global table_name_mm_deleted

                conn = get_db_connection()
                cursor = conn.cursor()

                data = request.get_json()
                cal_id = data.get('cal_id') 
                print(cal_id)

                # First insert into deleted table with specific columns
                insert_query = f"""
                INSERT INTO {table_name_mm_deleted} 
                (CALid, CALLoggerid, CALLogger, MarketCategory, DealReference, DealType, 
                Counterparty, Currency, EffectedDate, ValueDate, DaysDelay, 
                KnockoffCALid, SupposedApprover, Approver)
                SELECT 
                CALid, CALLoggerid, CALLogger, MarketCategory, DealReference, DealType, 
                Counterparty, Currency, EffectedDate, ValueDate, DaysDelay, 
                KnockoffCALid, SupposedApprover, Approver
                FROM {table_name_mm} 
                WHERE CALid = %s;
                """
                cursor.execute(insert_query, (cal_id,))
                
                # Then delete from main table
                delete_query = f"""
                DELETE FROM {table_name_mm} WHERE CALid = %s;
                """
                cursor.execute(delete_query, (cal_id,))

                conn.commit()  # Don't forget to commit the transaction!
                
                return jsonify({'success': True, 'message': 'CAL deleted successfully'})
            except Exception as e:
                if conn:
                    conn.rollback()  # Rollback on error
                return jsonify({'success': False, 'message': str(e)}), 500
            finally:
                if conn:
                    conn.close()  # Always close the connection
    

        @app.route('/excel_paste_import', methods=['POST'])
        def excel_paste_import():
            global table_name_mm, today_date
            tsv_data = request.form['tableData']


              
            try:
                conn = get_db_connection()
                cursor = conn.cursor()

                df = pd.read_csv(StringIO(tsv_data), sep='\t', header=None)
                df = df.dropna(how='all').dropna(axis=1, how='all')
                print(df)

                # Convert to datetime and filter
                df[1] = pd.to_datetime(df[1], errors='coerce')  # Column index 1
                df[2] = pd.to_datetime(df[2], errors='coerce')  # Column index 2
                filtered_df = df[df[2] > df[1]].copy()
                filtered_df = filtered_df.dropna(subset=[1, 2], how='any')
                print(filtered_df)

                query = f"SELECT DealReference FROM {table_name_mm};"
                cursor.execute(query)
                rows = cursor.fetchall()

                mmlog = pd.DataFrame(rows, columns=["DEAL REFERENCE"])
                print(mmlog)


                cursor.close()
                conn.close()

                values_to_remove = set(mmlog['DEAL REFERENCE'])

                filtered_df = filtered_df[~filtered_df.iloc[:, 0].isin(values_to_remove)]
                print(filtered_df)

                total_rows = len(filtered_df)
                success_count = 0
                error_rows = []

                for index, row in filtered_df.iterrows():
                    try:
                        # Extract values from columns
                        deal_reference = str(row[0]) if pd.notna(row[0]) else None
                        backdatedcapturevaluedate = row[1]  # Already a Timestamp
                        backdated_capture_date = row[2]  # Already a Timestamp

                        # Skip if any required field is empty
                        if not all([deal_reference, backdatedcapturevaluedate, backdated_capture_date]):
                            error_rows.append(f"Row {index+1}: Missing required fields")
                            continue
                        
                        # Convert Timestamps to strings in YYYY-MM-DD format
                        try:
                            backdatedcapturevaluedate_str = backdatedcapturevaluedate.strftime('%Y-%m-%d')
                            backdated_capture_date_str = backdated_capture_date.strftime('%Y-%m-%d')
                        except Exception as e:
                            error_rows.append(f"Row {index+1}: Date format error - {str(e)}")
                            continue
                        
                        # Process deal reference
                        deal_type_mapping = {
                            'nncd': "NNCD",
                            'inbr': "Treasury Bill",
                            'inop': "Offshore Placement",
                            'inpp': "Local Placement",
                            'fixd': "Fixed Deposit",
                            'inbd': "Interbank Deposit",
                            'cctd': "Cash Cover Term Deposit",
                            'pnot': "Promissory Note",
                            'fixp': "Fixed Treasury Placement",
                        }
                        
                        backdated_deal_type = deal_reference[:4].lower()
                        deal_type = deal_type_mapping.get(backdated_deal_type, "Unknown Deal Type")
                        
                        # Determine currency
                        if "za" in deal_reference.lower():
                            currency = "ZAR"
                        elif "eu" in deal_reference.lower():
                            currency = "EUR"
                        elif "zg" in deal_reference.lower():
                            currency = "ZWG"
                        else:
                            currency = "USD"
                        
                        # Calculate working days
                        try:
                            working_days_count = 0
                            current_date = backdatedcapturevaluedate
                            while current_date <= backdated_capture_date:
                                if current_date.weekday() < 5:  # Monday-Friday
                                    working_days_count += 1
                                current_date += timedelta(days=1)
                        except Exception as e:
                            error_rows.append(f"Row {index+1}: Date calculation error - {str(e)}")
                            continue
                        
                        # Database insertion
                        try:
                            conn = get_db_connection()
                            cursor = conn.cursor()

                            insert_query = f"""
                            INSERT INTO {table_name_mm} 
                            (DateLogged, CALLoggerid, CALLogger, Market, MarketCategory, 
                            DealReference, DealType, Currency, EffectedDate, ValueDate, 
                            DaysDelay, comments, count)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            
                            cursor.execute(insert_query, (
                                today_date,
                                session.get('empid'),
                                session.get('username'),
                                'Money Market',
                                'Backdated Capture',
                                deal_reference,
                                deal_type,
                                currency,
                                backdated_capture_date_str,
                                backdatedcapturevaluedate_str,
                                working_days_count - 1,
                                "Imported via Excel paste",
                                1
                            ))
                            
                            conn.commit()
                            success_count += 1
                            
                        except Exception as e:
                            error_rows.append(f"Row {index+1}: Database error - {str(e)}")
                            if 'conn' in locals():
                                conn.rollback()
                        finally:
                            if 'conn' in locals():
                                conn.close()
                                
                    except Exception as e:
                        error_rows.append(f"Row {index+1}: Processing error - {str(e)}")
                
                # Prepare JavaScript response
                js_script = f"""
                <script>
                    alert("Processing complete!\\n\\nSuccessfully processed {success_count} of {total_rows} rows");
                """
                
                if error_rows:
                    error_str = "\\n\\n".join(error_rows)
                    js_script += f"""
                    alert(`The following errors occurred:\\n\\n{error_str}`);
                    """
                
                js_script += """
                    window.location.href = '/dashboard';
                </script>
                """
                
                return js_script
                
            except Exception as e:
                return f"""
                <script>
                    alert("Fatal error during processing: {str(e)}");
                    window.location.href = '/dashboard';
                </script>
                """


        @app.route('/dashboard')
        def dashboard():

            user_uuid = session.get('user_uuid')
            if user_uuid:

                username = session.get('username')
                empid = session.get('empid')
                results = run1(empid)  

                return render_template('main.html', **results, username = username)

            
            else:
                return redirect(url_for('landingpage')) 

        @app.route('/login', methods=['POST'])
        def login():
            """User login route"""
            if request.method == 'POST':
                username = request.form.get('usernamelogin', '').strip()
                password = request.form.get('passwordlogin', '').strip()

                print(username)
                print(password)

                if not username or not password:
                    return jsonify({'success': False, 'message': 'Username and password are required.'}), 400

                conn = get_db_connection()
                if not conn:
                    return jsonify({'success': False, 'message': 'Database connection failed.'}), 500

                try:
                    cursor = conn.cursor()
                    query = f"SELECT id, username, password FROM {TABLE_NAME_USERS} WHERE username = %s;"
                    cursor.execute(query, (username,))
                    result = cursor.fetchone()
                    print(result)

                    if result:

                        db_id, db_username, db_password = result
                        user_uuid = uuid.uuid4()
                        session['user_uuid'] = str(user_uuid)
                        session.permanent = True
                        user_sessions[username] = {'uuid': str(user_uuid), 'username': db_username}
                        session['empid'] = int(np.int64(db_id)) 
                        session['username'] = db_username


                        if password == db_password:
                            return redirect(url_for('dashboard'))
                        else:
                            return jsonify({'success': False, 'message': 'Invalid credentials.'}), 401
                    else:
                        return jsonify({'success': False, 'message': 'User not found.'}), 404
                except Exception as e:
                    print("🚨 Error during login:", e)
                    return jsonify({'success': False, 'message': str(e)}), 500
                finally:
                    cursor.close()
                    release_db_connection(conn)

        @app.route('/logout')
        def logout():
            """Log out and clear session"""
            session.clear()
            return redirect(url_for('landingpage'))

        @app.route('/')
        def landingpage():
            return render_template('login.html')







except Exception as e:
    print("🚨 Error creating database connection pool:", e)
    exit()


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)  # Port changed from 5432 (Postgres) to 5000 (Flask default)
